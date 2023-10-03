import openpyxl
import json

from lib.product import Product

class ProductsContainer(object):
    def __init__(self, dict = None) -> None:
        self.products = []
        if dict:
            for product_dict in dict["products"]:
                p = Product(**product_dict)
                self.products.append(p)

    def append(self, product):
        self.products.append(product)

    def save_to_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        index = 1
        sheet.cell(row=index, column=2).value = "URI"
        sheet.cell(row=index, column=3).value = "Text"
        sheet.cell(row=index, column=4).value = "Price"
        sheet.cell(row=index, column=5).value = "iPhone"
        sheet.cell(row=index, column=6).value = "Battery"
        sheet.cell(row=index, column=7).value = "Apple Garantie"
        for index in range(2, len(self.products)+1):
            p = self.products[index-1]
            sheet.cell(row=index, column=2).value = p.link
            sheet.cell(row=index, column=3).value = p.description
            sheet.cell(row=index, column=4).value = p.price

        workbook.save("usedproducts.xlsx")

    def fill_derived(self):
        for product in self.products:
            product.fill_derived()
            
    def to_json(self):
        return json.dumps(self, indent = 4, default=lambda o: o.__dict__, ensure_ascii=False)

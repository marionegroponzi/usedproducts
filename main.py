#!/usr/bin/env python
import argparse
import os
import openpyxl
import urllib3
import logging
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from bs4 import BeautifulSoup, Tag, ResultSet
from dataclasses import dataclass
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import re
import math

class PageLoaded(object):
    def __call__(self, driver):
        r = driver.execute_script("return document.readyState")
        return r == "complete"




class Scanner(object):

    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.index = 1


        # options = selenium.webdriver.ChromeOptions()
        # options.set_capability('acceptInsecureCerts', True)
        # options.add_argument('--disable-notifications')

        # options.add_argument('--auth-server-whitelist=*')
        # options.add_argument('--disable-gpu')
        # options.add_argument('--headless')
        service = Service(ChromeDriverManager().install())
        options = webdriver.ChromeOptions()
        options.set_capability('acceptInsecureCerts', True)
        options.add_argument('--disable-notifications')

        self.browser = webdriver.Chrome(service=service, options=options)
        self.browser.set_page_load_timeout(15)

    def __del__(self):
        if hasattr(self, 'browser') and self.browser is not None:
            self.browser.quit()

    def close(self):
        if hasattr(self, 'browser') and self.browser is not None:
            self.browser.quit()

    def get_num_pages(self, uri):
        self.browser.get(uri)
        wait = WebDriverWait(self.browser, 10)
        wait.until(PageLoaded())
        content = self.browser.page_source
        soup = BeautifulSoup(content, "html.parser")

        text = soup.find(class_="woocommerce-result-count").text
        # text is expected to be in the form "        1–12 van de 28125 producten"
        matches = re.findall('\d+', text)
        num_pages = math.ceil(int(matches[2]) / int(matches[1]))
        return num_pages
    
    def scan(self, uri):
        self.browser.get(uri)
        wait = WebDriverWait(self.browser, 10)
        wait.until(PageLoaded())
        content = self.browser.page_source
        soup = BeautifulSoup(content, "html.parser")

        hs = soup.find_all(class_="product-thumnail")

        for h in hs:
            product = h.find(class_="product-name-collection")
            product_description = product.a.text
            product_link = product.a.get('href')
            product_price = h.find(class_="product-price").div.span.text
            print(f"{product_link} - {product_description} >> {product_price}")
            self.sheet.cell(row=self.index, column=2).value = product_link
            self.sheet.cell(row=self.index, column=3).value = product_description
            self.sheet.cell(row=self.index, column=4).value = product_price
            is_iphone = re.match("iphone", product_description, flags=re.I) is not None
            battery_level = re.search("accu (\d+)", product_description, flags=re.I)
            if battery_level:
                battery_level = battery_level[1]
            has_apple_garantie = re.match("apple garantie", product_description, flags=re.I) is not None
            print(f"    iPhone: {is_iphone} - Battery: {battery_level} - Apple Garantie: {has_apple_garantie}")
            self.sheet.cell(row=self.index, column=5).value = is_iphone
            self.sheet.cell(row=self.index, column=6).value = battery_level
            self.sheet.cell(row=self.index, column=7).value = has_apple_garantie

            self.index += 1


        self.workbook.save("usedproducts.xlsx")

def main():
    if not parse_args():
        return
    config_env()

    print('>>>>>>>>>> usedproducts start <<<<<<<<<<')
    scanner = Scanner()
    num_pages = scanner.get_num_pages("https://www.usedproducts.nl/page/1/?s&post_type=product&vestiging=0")
    # for page in range(1,num_pages):
    for page in range(1,10):
        scanner.scan(f"https://www.usedproducts.nl/page/{page}/?s&post_type=product&vestiging=0")
    scanner.close()
    print('<<<<<<<<<< usedproducts stop >>>>>>>>>>')


def parse_args():
    parser = argparse.ArgumentParser(description='Check dead links')
    parser.add_argument('--log', type=str,
                        help='loglevel: DEBUG, INFO, WARN, ERROR (default: ERROR)')
    args = parser.parse_args()
    if args.log is not None:
        numeric_level = getattr(logging, args.log.upper())
        if not isinstance(numeric_level, int):
            raise ValueError('Invalid log level: %s' % args.loglevel)
        logging.basicConfig(level=numeric_level)
    return True


def config_env():
    # some pages are self-signed by ABN AMRO, and we don't want to see all the warnings about those
    urllib3.disable_warnings()

    os.environ['PATH'] = f"{os.environ['PATH']}:{os.getcwd()}"


if __name__ == "__main__":
    main()
